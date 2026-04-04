from flask import Flask, request, jsonify, send_file, render_template
import random, time, tracemalloc, copy, io, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__,
    template_folder=os.path.join(os.path.dirname(__file__), '..', 'templates'),
    static_folder=os.path.join(os.path.dirname(__file__), '..', 'static'))

class Task:
    def __init__(self, name, deadline, priority, duration, weight):
        self.name = name; self.deadline = int(deadline)
        self.priority = int(priority); self.duration = float(duration)
        self.weight = float(weight); self.score = 0.0
        self.start_time = 0.0; self.end_time = 0.0; self.is_late = False
    def to_dict(self):
        return {"name": self.name, "deadline": self.deadline, "priority": self.priority,
                "duration": self.duration, "weight": self.weight, "score": round(self.score, 4),
                "start_time": round(self.start_time, 2), "end_time": round(self.end_time, 2),
                "is_late": self.is_late}

def compute_score(t, a=0.4, b=0.4, g=0.2):
    return a * (1/t.deadline) + b * t.priority + g * (t.weight/t.duration)

def greedy_scheduler(tasks, a=0.4, b=0.4, g=0.2):
    for t in tasks: t.score = compute_score(t, a, b, g)
    sorted_tasks = sorted(tasks, key=lambda t: t.score, reverse=True)
    ct = 0.0; schedule = []
    for t in sorted_tasks:
        t.start_time = ct; t.end_time = ct + t.duration
        t.is_late = t.end_time > (t.deadline * 8); ct += t.duration; schedule.append(t)
    return schedule

def random_scheduler(tasks):
    shuffled = tasks[:]; random.shuffle(shuffled); ct = 0.0
    for t in shuffled:
        t.start_time = ct; t.end_time = ct + t.duration
        t.is_late = t.end_time > (t.deadline * 8); ct += t.duration
    return shuffled

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/schedule', methods=['POST'])
def schedule():
    try:
        data = request.get_json()
        tasks_raw = data.get('tasks', [])
        if not tasks_raw:
            return jsonify({"error": "Tidak ada tugas."}), 400
        tasks = [Task(t['name'], t['deadline'], t['priority'], t['duration'], t['weight']) for t in tasks_raw]
        tasks_random = copy.deepcopy(tasks)
        tracemalloc.start()
        start = time.perf_counter()
        greedy_result = greedy_scheduler(tasks)
        elapsed = (time.perf_counter() - start) * 1000
        _, peak_mem = tracemalloc.get_traced_memory(); tracemalloc.stop()
        random_result = random_scheduler(tasks_random)
        lg = sum(1 for t in greedy_result if t.is_late)
        lr = sum(1 for t in random_result if t.is_late)
        reduction = round((lr - lg) / lr * 100, 1) if lr > 0 else 100.0
        return jsonify({
            "greedy": [t.to_dict() for t in greedy_result],
            "random": [t.to_dict() for t in random_result],
            "summary": {"total": len(greedy_result), "late_greedy": lg, "late_random": lr,
                        "reduction": reduction, "exec_time_ms": round(elapsed, 4),
                        "memory_kb": round(peak_mem / 1024, 3),
                        "tardiness": round(sum(t.end_time - t.deadline*8 for t in greedy_result if t.is_late), 2)}
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/download', methods=['POST'])
def download():
    try:
        data    = request.get_json()
        greedy  = data.get('greedy', [])
        rand    = data.get('random', [])
        summary = data.get('summary', {})

        wb = Workbook()

        # ── Style helpers ──
        def hdr_font(bold=True, color="FFFFFF", size=11):
            return Font(bold=bold, color=color, size=size, name="Calibri")
        def cell_font(bold=False, color="000000", size=10):
            return Font(bold=bold, color=color, size=size, name="Calibri")
        def fill(hex_color):
            return PatternFill("solid", fgColor=hex_color)
        def border():
            s = Side(style="thin", color="CCCCCC")
            return Border(left=s, right=s, top=s, bottom=s)
        def center():
            return Alignment(horizontal="center", vertical="center", wrap_text=True)

        priority_label = {1: "Rendah", 2: "Sedang", 3: "Tinggi"}

        def write_schedule_sheet(ws, schedule, title):
            # Title row
            ws.merge_cells("A1:J1")
            ws["A1"] = title
            ws["A1"].font = Font(bold=True, size=13, color="FFFFFF", name="Calibri")
            ws["A1"].fill = fill("1F4E79")
            ws["A1"].alignment = center()
            ws.row_dimensions[1].height = 28

            # Sub-title info
            ws.merge_cells("A2:J2")
            ws["A2"] = "Greedy Scoring Function · Strategi Algoritma 2025/2026 · Muthia Febrahma Khoirunnisa & Aam Aminah"
            ws["A2"].font = Font(italic=True, size=9, color="595959", name="Calibri")
            ws["A2"].alignment = Alignment(horizontal="center")

            # Header row
            headers = ["No", "Nama Tugas", "Skor", "Deadline\n(hari)", "Prioritas",
                       "Durasi\n(jam)", "Bobot\n(%)", "Mulai\n(jam)", "Selesai\n(jam)", "Status"]
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=4, column=col, value=h)
                c.font = hdr_font()
                c.fill = fill("2E74B5")
                c.alignment = center()
                c.border = border()
            ws.row_dimensions[4].height = 30

            # Data rows
            for i, t in enumerate(schedule, 1):
                row = i + 4
                is_late = t['is_late']
                row_fill = fill("FFF2F2") if is_late else (fill("F0FFF8") if i % 2 == 0 else fill("FFFFFF"))
                values = [i, t['name'], t['score'], t['deadline'],
                          priority_label.get(t['priority'], t['priority']),
                          t['duration'], t['weight'], t['start_time'], t['end_time'],
                          "TERLAMBAT" if is_late else "Tepat Waktu"]
                for col, val in enumerate(values, 1):
                    c = ws.cell(row=row, column=col, value=val)
                    c.fill = row_fill
                    c.border = border()
                    c.alignment = Alignment(horizontal="center", vertical="center")
                    c.font = cell_font(color="C00000" if is_late and col == 10 else "217346" if col == 10 else "000000",
                                       bold=(col == 10))
                ws.row_dimensions[row].height = 18

            # Column widths
            widths = [5, 28, 8, 10, 10, 8, 8, 10, 10, 14]
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

        # ── Sheet 1: Greedy ──
        ws1 = wb.active
        ws1.title = "Jadwal Greedy"
        write_schedule_sheet(ws1, greedy, "HASIL JADWAL GREEDY MULTI-KRITERIA")

        # ── Sheet 2: Random ──
        ws2 = wb.create_sheet("Jadwal Random")
        write_schedule_sheet(ws2, rand, "JADWAL ACAK (Baseline Pembanding)")

        # ── Sheet 3: Summary ──
        ws3 = wb.create_sheet("Ringkasan Performa")
        ws3.merge_cells("A1:C1")
        ws3["A1"] = "RINGKASAN PERFORMA ALGORITMA"
        ws3["A1"].font = Font(bold=True, size=13, color="FFFFFF", name="Calibri")
        ws3["A1"].fill = fill("1F4E79")
        ws3["A1"].alignment = center()
        ws3.row_dimensions[1].height = 28

        rows_summary = [
            ("Metrik", "Nilai", "Keterangan"),
            ("Total Tugas", summary.get('total', ''), "Jumlah tugas yang dijadwalkan"),
            ("Tugas Terlambat (Greedy)", summary.get('late_greedy', ''), "Jumlah tugas melewati deadline"),
            ("Tugas Terlambat (Random)", summary.get('late_random', ''), "Baseline pembanding"),
            ("Reduksi Keterlambatan", f"{summary.get('reduction', '')}%", "Efektivitas algoritma greedy"),
            ("Waktu Eksekusi", f"{summary.get('exec_time_ms', '')} ms", "Durasi komputasi algoritma"),
            ("Memori (peak)", f"{summary.get('memory_kb', '')} KB", "Penggunaan memori puncak"),
            ("Total Tardiness", f"{summary.get('tardiness', '')} jam", "Total jam keterlambatan"),
            ("Time Complexity", "O(n log n)", "Dominated by sorting step"),
            ("Space Complexity", "O(n)", "Linear space untuk n tugas"),
            ("Complexity Class", "P (Polynomial)", "Diselesaikan secara deterministik"),
        ]
        for ri, row_data in enumerate(rows_summary, 2):
            for ci, val in enumerate(row_data, 1):
                c = ws3.cell(row=ri, column=ci, value=val)
                c.border = border()
                c.alignment = Alignment(horizontal="left", vertical="center")
                if ri == 2:
                    c.font = hdr_font()
                    c.fill = fill("2E74B5")
                    c.alignment = center()
                else:
                    c.font = cell_font(bold=(ci == 1))
                    c.fill = fill("EBF3FB") if ri % 2 == 0 else fill("FFFFFF")
            ws3.row_dimensions[ri].height = 18
        ws3.column_dimensions["A"].width = 28
        ws3.column_dimensions["B"].width = 20
        ws3.column_dimensions["C"].width = 38

        # Save to buffer
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        return send_file(buf,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True,
                         download_name='hasil_greedy_scoring_function.xlsx')

    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True)
